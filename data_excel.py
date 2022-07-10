from openpyxl import load_workbook


#Extracción de datos por medio de iterador primerafila...

Excel = "GESTION.xlsx"

# Conexión con Openpyxl
Wb = load_workbook(Excel, data_only=True)
sheet = Wb["Hoja1"]

initial_data = [""] * 2 # Se genera doble espacio vacio inicialmente...


def fila_excel():
    delete = [0,4,9,10] # Datos descartados
    for value in sheet.iter_rows(min_row = 1, max_row = 1,values_only=True):
            elementos = list(value)
            for i in sorted(delete, reverse = True):
                del elementos[i]

    elementos[1] = elementos[1].strftime("%Y%m%d")# cambio de formato de date time a string
    elementos[3] = elementos[3].strftime("%Y%m%d")
    elementos = initial_data + elementos  # Arrays concatenados para confrontación de fila excel con coordenadas...

    return elementos 

